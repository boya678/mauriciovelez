"""
Contactos API  (admin only)

GET /contactos          — paginated list of contacts with tags
GET /contactos/export   — Excel file with all contacts, tags as columns
"""
import io
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.tenant import get_tenant_db, require_admin, resolve_tenant, TenantContext

router = APIRouter(prefix="/contactos", tags=["contactos"])


class ContactOut(BaseModel):
    phone: str
    tags: str


def _parse_tags(raw: str) -> dict[str, str]:
    """Parse 'key:value,key2:value2' into {key: value}."""
    result: dict[str, str] = {}
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        colon = part.find(":")
        if colon >= 0:
            result[part[:colon].strip()] = part[colon + 1:].strip()
        else:
            result[part] = ""
    return result


@router.get("/export")
async def export_contacts_excel(
    tenant: TenantContext = Depends(resolve_tenant),
    db: AsyncSession = Depends(get_tenant_db),
    _admin=Depends(require_admin),
):
    schema = tenant.schema
    rows = (await db.execute(
        text(f"SELECT id, tags FROM {schema}.contactos ORDER BY id")
    )).fetchall()

    # Parse all tags and collect unique keys (columns)
    parsed = [(r[0], _parse_tags(r[1] or "")) for r in rows]
    key_set: set[str] = set()
    for _, tags in parsed:
        key_set.update(tags.keys())
    keys = sorted(key_set)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"

    # Header row
    headers = ["Teléfono"] + keys
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows — all values as text to avoid Excel reformatting
    for phone, tags in parsed:
        row = [str(phone)] + [str(tags.get(k, "")) for k in keys]
        ws.append(row)
        # Force every cell in this row to text format
        for cell in ws[ws.max_row]:
            cell.number_format = "@"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="contactos.xlsx"'},
    )


@router.get("", response_model=list[ContactOut])
async def list_contacts(
    search: Optional[str] = Query(None, description="Filtra por teléfono o tag"),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=1000),
    tenant: TenantContext = Depends(resolve_tenant),
    db: AsyncSession = Depends(get_tenant_db),
    _admin=Depends(require_admin),
):
    schema = tenant.schema
    offset = (page - 1) * page_size

    if search:
        sql = text(
            f"SELECT id, tags FROM {schema}.contactos "
            f"WHERE id ILIKE :q OR tags ILIKE :q "
            f"ORDER BY id LIMIT :limit OFFSET :offset"
        )
        rows = await db.execute(sql, {"q": f"%{search}%", "limit": page_size, "offset": offset})
    else:
        sql = text(
            f"SELECT id, tags FROM {schema}.contactos "
            f"ORDER BY id LIMIT :limit OFFSET :offset"
        )
        rows = await db.execute(sql, {"limit": page_size, "offset": offset})

    return [ContactOut(phone=r[0], tags=r[1] or "") for r in rows.fetchall()]
