import openpyxl
from datetime import timezone
import psycopg2
import re

DATABASE_URL = 'postgresql://postgres:Ardilla1*@transferiadb.postgres.database.azure.com:5432/portal'

m = re.match(r'postgresql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', DATABASE_URL)
user, password, host, port, dbname = m.groups()

conn = psycopg2.connect(host=host, port=int(port), dbname=dbname, user=user, password=password, sslmode='require')
cur = conn.cursor()

EXCEL_PATH = r'c:\Users\boya6\OneDrive\Documents\maxibingo\mauriciovelez\CONTROL VIP (1).xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

COL_ACTIVA  = 10  # J
COL_FECHA   = 11  # K
COL_SUS_ID  = 5   # E

ws.cell(row=1, column=COL_ACTIVA).value  = "Suscripción activa"
ws.cell(row=1, column=COL_FECHA).value   = "Fecha vencimiento"
ws.cell(row=1, column=COL_SUS_ID).value  = "Último ID suscripción"

procesados = activas = sin_activa = sin_celular = 0

for row in ws.iter_rows(min_row=2):
    celular_raw = row[1].value  # columna B

    if not celular_raw:
        ws.cell(row=row[0].row, column=COL_ACTIVA).value  = "Sin celular"
        ws.cell(row=row[0].row, column=COL_FECHA).value   = ""
        ws.cell(row=row[0].row, column=COL_SUS_ID).value  = ""
        sin_celular += 1
        continue

    # Quitar código de país 57 y dejar los 10 dígitos
    celular_str = str(int(celular_raw)) if isinstance(celular_raw, float) else str(celular_raw).strip()
    if celular_str.startswith('57') and len(celular_str) == 12:
        celular_str = celular_str[2:]

    # Buscar la última suscripción (activa o no) para obtener su ID
    cur.execute("""
        SELECT s.id, s.fin, s.activa FROM suscripciones s
        JOIN clientes c ON s.cliente_id = c.id
        WHERE c.celular = %s
        ORDER BY s.fin DESC LIMIT 1
    """, (celular_str,))
    ultima_row = cur.fetchone()

    if ultima_row:
        sus_id, fin, es_activa = ultima_row
        if fin.tzinfo:
            fin = fin.astimezone(timezone.utc).replace(tzinfo=None)
        ws.cell(row=row[0].row, column=COL_SUS_ID).value = str(sus_id)
        ws.cell(row=row[0].row, column=COL_FECHA).value  = fin.strftime('%d/%m/%Y')
        if es_activa:
            ws.cell(row=row[0].row, column=COL_ACTIVA).value = "Sí"
            activas += 1
        else:
            ws.cell(row=row[0].row, column=COL_ACTIVA).value = "No"
            sin_activa += 1
    else:
        ws.cell(row=row[0].row, column=COL_ACTIVA).value = "No"
        ws.cell(row=row[0].row, column=COL_FECHA).value  = "Sin registro"
        ws.cell(row=row[0].row, column=COL_SUS_ID).value = ""
        sin_activa += 1

    procesados += 1
    if procesados % 100 == 0:
        print(f"  {procesados} procesados...")

cur.close()
conn.close()

wb.save(EXCEL_PATH)
print(f"\n✔  Listo — {procesados} filas procesadas")
print(f"   Con suscripción activa : {activas}")
print(f"   Sin suscripción activa : {sin_activa}")
print(f"   Sin celular            : {sin_celular}")


DATABASE_URL = 'postgresql://postgres:Ardilla1*@transferiadb.postgres.database.azure.com:5432/portal'

m = re.match(r'postgresql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', DATABASE_URL)
user, password, host, port, dbname = m.groups()

conn = psycopg2.connect(host=host, port=int(port), dbname=dbname, user=user, password=password, sslmode='require')
cur = conn.cursor()

EXCEL_PATH = r'c:\Users\boya6\OneDrive\Documents\maxibingo\mauriciovelez\CONTROL VIP (1).xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Columna J = celular encontrado en BD, K = suscripción activa
COL_CELULAR = 10  # J
COL_ACTIVA  = 11  # K

ws.cell(row=1, column=COL_CELULAR).value = "Celular en BD"
ws.cell(row=1, column=COL_ACTIVA).value  = "Suscripción activa"

procesados = activas = sin_activa = sin_celular = 0

for row in ws.iter_rows(min_row=2):
    mes2    = row[3].value  # columna D
    if mes2 is None:
        continue

    celular = row[1].value  # columna B
    if not celular:
        ws.cell(row=row[0].row, column=COL_CELULAR).value = "Sin celular"
        ws.cell(row=row[0].row, column=COL_ACTIVA).value  = "-"
        sin_celular += 1
        continue

    celular_str = str(int(celular)) if isinstance(celular, float) else str(celular).strip()

    cur.execute('''
        SELECT EXISTS (
            SELECT 1 FROM suscripciones s
            JOIN clientes c ON s.cliente_id = c.id
            WHERE c.celular = %s AND s.activa = true
        )
    ''', (celular_str,))
    tiene_activa = cur.fetchone()[0]

    ws.cell(row=row[0].row, column=COL_CELULAR).value = celular_str
    ws.cell(row=row[0].row, column=COL_ACTIVA).value  = "Sí" if tiene_activa else "No"

    if tiene_activa:
        activas += 1
    else:
        sin_activa += 1

    procesados += 1
    if procesados % 50 == 0:
        print(f"  {procesados} procesados...")

cur.close()
conn.close()

wb.save(EXCEL_PATH)
print(f"\n✔  Listo — {procesados} filas con Mes 2 procesadas")
print(f"   Con suscripción activa : {activas}")
print(f"   Sin suscripción activa : {sin_activa}")
print(f"   Sin celular            : {sin_celular}")


DATABASE_URL = 'postgresql://postgres:Ardilla1*@transferiadb.postgres.database.azure.com:5432/portal'

m = re.match(r'postgresql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', DATABASE_URL)
user, password, host, port, dbname = m.groups()

conn = psycopg2.connect(host=host, port=int(port), dbname=dbname, user=user, password=password, sslmode='require')
cur = conn.cursor()

EXCEL_PATH = r'c:\Users\boya6\OneDrive\Documents\maxibingo\mauriciovelez\CONTROL VIP (1).xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Columnas nuevas: J = "Suscripción activa (fin)", K = "Vencida el"
COL_ACTIVA  = 10  # J
COL_VENCIDA = 11  # K

# Cabeceras en fila 1
ws.cell(row=1, column=COL_ACTIVA).value  = "Suscripción activa (vence)"
ws.cell(row=1, column=COL_VENCIDA).value = "Vencida el"

procesados = 0
activas = 0
sin_activa = 0

for i, row in enumerate(ws.iter_rows(min_row=2)):
    celular = row[1].value  # columna B
    mes2    = row[3].value  # columna D

    if mes2 is None or not celular:
        continue

    celular_str = str(int(celular)) if isinstance(celular, float) else str(celular).strip()

    # 1. Buscar suscripción ACTIVA
    cur.execute('''
        SELECT s.fin
        FROM suscripciones s
        JOIN clientes c ON s.cliente_id = c.id
        WHERE c.celular = %s AND s.activa = true
        ORDER BY s.fin DESC
        LIMIT 1
    ''', (celular_str,))
    activa_row = cur.fetchone()

    if activa_row:
        fin = activa_row[0]
        if fin.tzinfo:
            fin = fin.astimezone(timezone.utc).replace(tzinfo=None)
        ws.cell(row=row[0].row, column=COL_ACTIVA).value  = fin.strftime('%d/%m/%Y')
        ws.cell(row=row[0].row, column=COL_VENCIDA).value = ""
        activas += 1
    else:
        # 2. Buscar la última suscripción inactiva (para saber cuándo venció)
        cur.execute('''
            SELECT s.fin
            FROM suscripciones s
            JOIN clientes c ON s.cliente_id = c.id
            WHERE c.celular = %s AND s.activa = false
            ORDER BY s.fin DESC
            LIMIT 1
        ''', (celular_str,))
        inactiva_row = cur.fetchone()

        ws.cell(row=row[0].row, column=COL_ACTIVA).value = "No"
        if inactiva_row:
            fin_venc = inactiva_row[0]
            if fin_venc.tzinfo:
                fin_venc = fin_venc.astimezone(timezone.utc).replace(tzinfo=None)
            ws.cell(row=row[0].row, column=COL_VENCIDA).value = fin_venc.strftime('%d/%m/%Y')
        else:
            ws.cell(row=row[0].row, column=COL_VENCIDA).value = "Sin registro"
        sin_activa += 1

    procesados += 1
    if procesados % 50 == 0:
        print(f"  {procesados} procesados...")

cur.close()
conn.close()

wb.save(EXCEL_PATH)

print(f"\n✔  Listo — {procesados} filas con Mes 2 procesadas")
print(f"   Con suscripción activa : {activas}")
print(f"   Sin suscripción activa : {sin_activa}")
print(f"   Archivo guardado: {EXCEL_PATH}")
