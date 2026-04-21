import openpyxl
import psycopg2
import re

DATABASE_URL = 'postgresql://postgres:Ardilla1*@transferiadb.postgres.database.azure.com:5432/portal'

m = re.match(r'postgresql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', DATABASE_URL)
user, password, host, port, dbname = m.groups()

conn = psycopg2.connect(host=host, port=int(port), dbname=dbname, user=user, password=password, sslmode='require')
cur = conn.cursor()

EXCEL_PATH = r'c:\Users\boya6\OneDrive\Documents\maxibingo\mauriciovelez\CONTROL VIP (1).xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

COL_SUS_ID = 5  # E

ws.cell(row=1, column=COL_SUS_ID).value = "Último ID suscripción"

procesados = encontrados = sin_registro = sin_celular = 0

for row in ws.iter_rows(min_row=2):
    celular_raw = row[1].value  # columna B

    if not celular_raw:
        ws.cell(row=row[0].row, column=COL_SUS_ID).value = ""
        sin_celular += 1
        continue

    celular_str = str(int(celular_raw)) if isinstance(celular_raw, float) else str(celular_raw).strip()
    if celular_str.startswith('57') and len(celular_str) == 12:
        celular_str = celular_str[2:]

    cur.execute("""
        SELECT s.id FROM suscripciones s
        JOIN clientes c ON s.cliente_id = c.id
        WHERE c.celular = %s AND s.activa = true
        ORDER BY s.fin DESC LIMIT 1
    """, (celular_str,))
    row_db = cur.fetchone()

    if row_db:
        ws.cell(row=row[0].row, column=COL_SUS_ID).value = str(row_db[0])
        encontrados += 1
    else:
        ws.cell(row=row[0].row, column=COL_SUS_ID).value = ""
        sin_registro += 1

    procesados += 1
    if procesados % 100 == 0:
        print(f"  {procesados} procesados...")

cur.close()
conn.close()

wb.save(EXCEL_PATH)
print(f"\n✔  Listo — {procesados} filas procesadas")
print(f"   Con suscripción encontrada : {encontrados}")
print(f"   Sin registro en BD         : {sin_registro}")
print(f"   Sin celular                : {sin_celular}")
