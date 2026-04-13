import asyncio
import json
import os
import threading

import aiohttp
import openpyxl
from dotenv import load_dotenv

load_dotenv()

WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
WHATSAPP_PHONE_ID = os.getenv("WHATSAPP_PHONE_ID")
API_URL = f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_ID}/messages"

EXCEL_FILE = "Whaticket y Software.xlsx"
BATCH_SIZE = 7000
MAX_CONCURRENT = 20   # peticiones simultáneas — sube o baja según límite de Meta
SAVE_EVERY = 50       # guarda Excel cada N resultados procesados

with open("body.json", "r", encoding="utf-8") as f:
    body_template = json.load(f)

headers_http = {
    "Authorization": f"Bearer {WHATSAPP_TOKEN}",
    "Content-Type": "application/json",
}

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.worksheets[0]

# Recolectar filas pendientes
pendientes: list[tuple] = []
for row in ws.iter_rows(min_row=2):
    if len(pendientes) >= BATCH_SIZE:
        break
    cell_numero = row[0]
    cell_estado = row[1] if len(row) > 1 else ws.cell(row=cell_numero.row, column=2)

    numero = cell_numero.value
    if not numero:
        continue
    if str(cell_estado.value).strip().upper() == "ENVIADO":
        continue

    numero_str = str(int(numero)) if isinstance(numero, float) else str(numero).strip()
    pendientes.append((numero_str, cell_estado))

print(f"Iniciando envío async — {len(pendientes)} pendientes | concurrencia: {MAX_CONCURRENT}\n")

# Estado compartido
lock = threading.Lock()
enviados = 0
errores = 0
procesados = 0


async def enviar(session: aiohttp.ClientSession, semaphore: asyncio.Semaphore,
                 numero_str: str, cell_estado) -> None:
    global enviados, errores, procesados

    body = json.loads(json.dumps(body_template))
    body["to"] = numero_str

    async with semaphore:
        try:
            async with session.post(
                API_URL, headers=headers_http, json=body,
                timeout=aiohttp.ClientTimeout(total=30)
            ) as resp:
                resp_json = await resp.json()

                if resp.status == 200 and "messages" in resp_json:
                    msg_id = resp_json["messages"][0].get("id", "?")
                    print(f"[OK]  {numero_str} | id: {msg_id}")
                    with lock:
                        cell_estado.value = "ENVIADO"
                        enviados += 1
                else:
                    print(f"[ERR] {numero_str} | HTTP {resp.status} | {resp_json}")
                    with lock:
                        errores += 1

        except Exception as e:
            print(f"[EXC] {numero_str} | {e}")
            with lock:
                errores += 1

        finally:
            with lock:
                procesados += 1
                if procesados % SAVE_EVERY == 0:
                    wb.save(EXCEL_FILE)
                    print(f"  → Excel guardado ({procesados} procesados)")


async def main():
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    async with aiohttp.ClientSession() as session:
        tareas = [
            enviar(session, semaphore, num, cell)
            for num, cell in pendientes
        ]
        await asyncio.gather(*tareas)

    wb.save(EXCEL_FILE)
    print(f"\n=== Listo: {enviados} enviados OK, {errores} errores ===")


asyncio.run(main())

print(f"\n=== Listo: {enviados} enviados OK, {errores} errores ===")
