"""
check_estado_entrega.py
Consulta el estado de entrega de mensajes WA ya enviados.

Lee el mismo Excel que enviar_difusion.py:
  col1 = número destino
  col3 = estado  (busca "ENVIADO")
  col4 = message_id
  col5 = estado entrega  (se escribe aquí)

Estados devueltos por la API:
  sent       → enviado al servidor de WA
  delivered  → entregado al dispositivo
  read       → leído
  failed     → falló
"""

import asyncio
import os
import threading

import aiohttp
import openpyxl
from dotenv import load_dotenv

load_dotenv()

WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Ganadores Numero Gratis 22042026 con loteria.xlsx")

MAX_CONCURRENT = 40
SAVE_EVERY     = 200

# Columna donde se escribe el estado de entrega (5 = E)
COL_ENTREGA = 5

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.worksheets[0]

# Agrega encabezado en col5 si no existe
if ws.cell(row=1, column=COL_ENTREGA).value is None:
    ws.cell(row=1, column=COL_ENTREGA).value = "estado_entrega"

# Recolecta filas con estado ENVIADO y message_id presente
pendientes: list[tuple] = []
for row in ws.iter_rows(min_row=2):
    cell_estado  = row[2] if len(row) > 2 else None
    cell_msgid   = row[3] if len(row) > 3 else None
    cell_entrega = ws.cell(row=row[0].row, column=COL_ENTREGA)

    if cell_estado is None or cell_msgid is None:
        continue

    estado = str(cell_estado.value or "").strip().upper()
    msg_id = str(cell_msgid.value or "").strip()

    if estado != "ENVIADO" or not msg_id:
        continue

    pendientes.append((msg_id, cell_entrega, row[0].value))

print(f"Mensajes a verificar: {len(pendientes)}\n")

lock       = threading.Lock()
verificados = 0
procesados  = 0

headers_http = {
    "Authorization": f"Bearer {WHATSAPP_TOKEN}",
}


async def consultar(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    msg_id: str,
    cell_entrega,
    numero,
) -> None:
    global verificados, procesados

    url = f"https://graph.facebook.com/v18.0/{msg_id}"
    params = {"fields": "id,status"}

    async with semaphore:
        try:
            async with session.get(
                url,
                headers=headers_http,
                params=params,
                timeout=aiohttp.ClientTimeout(total=20),
            ) as resp:
                data = await resp.json()

                if resp.status == 200:
                    status = data.get("status", "desconocido")
                    print(f"[OK]  {numero} | {msg_id} → {status}")
                    with lock:
                        cell_entrega.value = status
                        verificados += 1
                else:
                    error_msg = data.get("error", {}).get("message", str(data))
                    print(f"[ERR] {numero} | {msg_id} | HTTP {resp.status} | {error_msg}")
                    with lock:
                        cell_entrega.value = f"ERROR {resp.status}"

        except Exception as e:
            print(f"[EXC] {numero} | {msg_id} | {e}")
            with lock:
                cell_entrega.value = "ERROR EXCEPCION"

        finally:
            with lock:
                procesados += 1
                if procesados % SAVE_EVERY == 0:
                    wb.save(EXCEL_FILE)
                    print(f"  → Excel guardado ({procesados}/{len(pendientes)})")


async def main():
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    async with aiohttp.ClientSession() as session:
        tareas = [
            consultar(session, semaphore, mid, cel, num)
            for mid, cel, num in pendientes
        ]
        await asyncio.gather(*tareas)

    wb.save(EXCEL_FILE)

    no_leidos    = sum(1 for _, c, _ in pendientes if c.value == "sent")
    entregados   = sum(1 for _, c, _ in pendientes if c.value == "delivered")
    leidos       = sum(1 for _, c, _ in pendientes if c.value == "read")
    fallidos     = sum(1 for _, c, _ in pendientes if str(c.value or "").startswith("ERROR"))

    print(f"\n=== Resumen ===")
    print(f"  read:      {leidos}")
    print(f"  delivered: {entregados}")
    print(f"  sent:      {no_leidos}")
    print(f"  errores:   {fallidos}")
    print(f"  total:     {verificados} verificados de {len(pendientes)}")


asyncio.run(main())
