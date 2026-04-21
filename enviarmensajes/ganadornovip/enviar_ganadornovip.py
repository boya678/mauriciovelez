import asyncio
import json
import os
import threading

import aiohttp
import openpyxl
from dotenv import load_dotenv

# Carga .env desde la raíz del proyecto (un nivel arriba de enviarmensajes)
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", "..", "backend", ".env"))

WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
WHATSAPP_PHONE_ID = os.getenv("WHATSAPP_PHONE_ID")
API_URL = f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_ID}/messages"

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "ganadornovip 20042026.xlsx")
BODY_FILE = os.path.join(os.path.dirname(__file__), "body.json")

MAX_CONCURRENT = 60
SAVE_EVERY = 500

with open(BODY_FILE, "r", encoding="utf-8") as f:
    body_template = json.load(f)

headers_http = {
    "Authorization": f"Bearer {WHATSAPP_TOKEN}",
    "Content-Type": "application/json",
}

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.worksheets[0]

# Recolectar filas pendientes (columna 1 desde fila 2)
pendientes: list[tuple] = []
for row in ws.iter_rows(min_row=2):
    cell_numero = row[0]
    cell_estado = row[1] if len(row) > 1 else ws.cell(row=cell_numero.row, column=2)

    numero = cell_numero.value
    if not numero:
        continue
    if str(cell_estado.value).strip().upper() == "ENVIADO":
        continue

    numero_str = str(int(numero)) if isinstance(numero, float) else str(numero).strip()
    pendientes.append((numero_str, cell_estado))

print(f"Iniciando envío — {len(pendientes)} pendientes | concurrencia: {MAX_CONCURRENT}\n")

lock = threading.Lock()
enviados = 0
errores = 0
procesados = 0


async def enviar(session: aiohttp.ClientSession, semaphore: asyncio.Semaphore,
                 numero_str: str, cell_estado) -> None:
    global enviados, errores, procesados

    body = json.loads(json.dumps(body_template))
    body["to"] = numero_str

    async with semaphore:
        try:
            async with session.post(API_URL, json=body, headers=headers_http) as resp:
                with lock:
                    procesados += 1
                    if resp.status == 200:
                        enviados += 1
                        cell_estado.value = "ENVIADO"
                    else:
                        errores += 1
                        text = await resp.text()
                        cell_estado.value = f"ERROR {resp.status}"
                        print(f"  ✗ {numero_str} → {resp.status}: {text[:120]}")

                    if procesados % SAVE_EVERY == 0:
                        wb.save(EXCEL_FILE)
                        print(f"  [guardado] {procesados}/{len(pendientes)} — ✓{enviados} ✗{errores}")
        except Exception as e:
            with lock:
                procesados += 1
                errores += 1
                cell_estado.value = "ERROR EXCEPCION"
                print(f"  ✗ {numero_str} → excepción: {e}")


async def main():
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    async with aiohttp.ClientSession() as session:
        tareas = [
            enviar(session, semaphore, num, cel)
            for num, cel in pendientes
        ]
        await asyncio.gather(*tareas)

    wb.save(EXCEL_FILE)
    print(f"\nFinalizado — ✓ enviados: {enviados} | ✗ errores: {errores} | total: {procesados}")


if __name__ == "__main__":
    asyncio.run(main())
