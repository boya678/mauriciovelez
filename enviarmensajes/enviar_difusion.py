import asyncio
import json
import os
import threading

import aiohttp
import openpyxl
from dotenv import load_dotenv

load_dotenv()

WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
WHATSAPP_PHONE_ID = os.getenv("WHATSAPP_PHONE_ID")
API_URL = f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_ID}/messages"

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Ganadores numero gratis 30042026.xlsx")
BODY_FILE  = os.path.join(os.path.dirname(__file__), "body.json")

MAX_CONCURRENT = 60
SAVE_EVERY     = 500
LIMITE         = None      # None = enviar todo | N = enviar solo las primeras N filas (prueba)

with open(BODY_FILE, "r", encoding="utf-8") as f:
    body_template = json.load(f)

headers_http = {
    "Authorization": f"Bearer {WHATSAPP_TOKEN}",
    "Content-Type": "application/json",
}

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.worksheets[0]

# col1=número destino  col2=variable del body ("numero")  col3=estado  col4=message_id
pendientes: list[tuple] = []
for row in ws.iter_rows(min_row=2):
    cell_numero   = row[0]                                                      # col 1
    cell_variable = row[1] if len(row) > 1 else ws.cell(row=cell_numero.row, column=2)  # col 2
    cell_estado   = row[2] if len(row) > 2 else ws.cell(row=cell_numero.row, column=3)  # col 3
    cell_msgid    = row[3] if len(row) > 3 else ws.cell(row=cell_numero.row, column=4)  # col 4

    numero = cell_numero.value
    if not numero:
        continue
    if str(cell_estado.value).strip().upper() == "ENVIADO":
        continue

    numero_str   = str(int(numero)) if isinstance(numero, float) else str(numero).strip()
    variable_str = str(cell_variable.value).strip() if cell_variable.value is not None else ""
    pendientes.append((numero_str, variable_str, cell_estado, cell_msgid))
    if LIMITE is not None and len(pendientes) >= LIMITE:
        break

print(f"Iniciando envío — {len(pendientes)} pendientes | concurrencia: {MAX_CONCURRENT}\n")

lock      = threading.Lock()
enviados  = 0
errores   = 0
procesados = 0


def _reemplazar_variable(body: dict, valor: str) -> dict:
    """Reemplaza el primer parámetro de texto 'numero' en los components del body."""
    for component in body.get("template", {}).get("components", []):
        if component.get("type") == "body":
            for param in component.get("parameters", []):
                if param.get("type") == "text" and param.get("text") == "numero":
                    param["text"] = valor
    return body


async def enviar(session: aiohttp.ClientSession, semaphore: asyncio.Semaphore,
                 numero_str: str, variable_str: str, cell_estado, cell_msgid) -> None:
    global enviados, errores, procesados

    body = json.loads(json.dumps(body_template))
    body["to"] = numero_str
    body = _reemplazar_variable(body, variable_str)

    async with semaphore:
        try:
            async with session.post(
                API_URL, headers=headers_http, json=body,
                timeout=aiohttp.ClientTimeout(total=30)
            ) as resp:
                resp_json = await resp.json()

                if resp.status == 200 and "messages" in resp_json:
                    msg_id = resp_json["messages"][0].get("id", "?")
                    print(f"[OK]  {numero_str} | var: {variable_str} | id: {msg_id}")
                    with lock:
                        cell_estado.value = "ENVIADO"
                        cell_msgid.value  = msg_id
                        enviados += 1
                else:
                    print(f"[ERR] {numero_str} | HTTP {resp.status} | {resp_json}")
                    with lock:
                        cell_estado.value = f"ERROR {resp.status}"
                        errores += 1

        except Exception as e:
            print(f"[EXC] {numero_str} | {e}")
            with lock:
                cell_estado.value = "ERROR EXCEPCION"
                errores += 1

        finally:
            with lock:
                procesados += 1
                if procesados % SAVE_EVERY == 0:
                    wb.save(EXCEL_FILE)
                    print(f"  → Excel guardado ({procesados}/{len(pendientes)} procesados)")


async def main():
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    async with aiohttp.ClientSession() as session:
        tareas = [
            enviar(session, semaphore, num, var, cel, mid)
            for num, var, cel, mid in pendientes
        ]
        await asyncio.gather(*tareas)

    wb.save(EXCEL_FILE)
    print(f"\n=== Listo: {enviados} enviados OK, {errores} errores ===")


asyncio.run(main())
